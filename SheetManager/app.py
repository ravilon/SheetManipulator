import sys
import csv
import math
import os
import zipfile
import threading
from pathlib import Path
from tkinter import Tk, Label, Button, Entry, StringVar, IntVar, filedialog, messagebox, Frame, Radiobutton, ttk
from tkinter import font as tkfont

# Aumenta o limite para campos gigantes
csv.field_size_limit(sys.maxsize)

def detect_delimiter(csv_path, encoding="utf-8"):
    with open(csv_path, "r", encoding=encoding) as f:
        first_line = f.readline()
        if ";" in first_line and "," not in first_line: return ";"
        if "," in first_line and ";" not in first_line: return ","
        return ";" if first_line.count(";") > first_line.count(",") else ","

def is_excel(path):
    return Path(path).suffix.lower() in (".xlsx", ".xls")

def iter_excel_rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield list(row)
    wb.close()

def split_file(
    input_path, # Nome genérico para evitar o erro de 'input_csv'
    output_dir,
    output_format="xlsx",
    num_files=None,
    rows_per_file=None,
    encoding="utf-8",
    progress_callback=None,
):
    input_is_excel = is_excel(input_path)
    base_name = Path(input_path).stem
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 1. Contagem de linhas (se necessário)
    if num_files is not None:
        if progress_callback: progress_callback("Contando linhas...")
        if input_is_excel:
            total_rows = sum(1 for _ in iter_excel_rows(input_path)) - 1
        else:
            delim = detect_delimiter(input_path, encoding)
            with open(input_path, "r", encoding=encoding) as f:
                total_rows = sum(1 for _ in f) - 1
        rows_per_file = math.ceil(total_rows / num_files)

    # 2. Preparação do Iterador de Entrada
    if input_is_excel:
        row_iter = iter_excel_rows(input_path)
    else:
        delim = detect_delimiter(input_path, encoding)
        _f = open(input_path, "r", encoding=encoding, newline="")
        row_iter = csv.reader(_f, delimiter=delim)

    # 3. Processamento
    try:
        header = next(row_iter, None)
        created_files = []
        part_idx = 1
        
        while True:
            rows_to_write = []
            for _ in range(rows_per_file):
                try:
                    row = next(row_iter)
                    rows_to_write.append(row)
                except StopIteration:
                    break
            
            if not rows_to_write: break

            suffix = f"{part_idx:04d}"
            if output_format == "csv":
                out_path = output_dir / f"{base_name}_part{suffix}.csv"
                with open(out_path, "w", encoding=encoding, newline="") as f:
                    writer = csv.writer(f, delimiter=";")
                    writer.writerow(header)
                    writer.writerows(rows_to_write)
            else:
                from openpyxl import Workbook
                out_path = output_dir / f"{base_name}_part{suffix}.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.append(header)
                for r in rows_to_write: ws.append(r)
                wb.save(out_path)
            
            created_files.append(str(out_path))
            if progress_callback: progress_callback(f"Arquivo {part_idx} criado...")
            part_idx += 1
            
    finally:
        if not input_is_excel: _f.close()
        
    return created_files

class CSVSplitterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("SheetManager - Splitter")
        self.root.geometry("600x550")
        
        self.input_file = StringVar()
        self.output_dir = StringVar()
        self.split_mode = IntVar(value=1)
        self.num_files = IntVar(value=5)
        self.rows_per_file = IntVar(value=100000)
        self.output_format = StringVar(value="xlsx")
        
        # UI Layout (Simplificado para o exemplo)
        Label(root, text="📊 SheetManager Splitter", font=("Segoe UI", 16, "bold")).pack(pady=20)
        
        f1 = Frame(root); f1.pack(padx=20, pady=10, fill="x")
        Label(f1, text="Arquivo (CSV ou Excel):").grid(row=0, column=0)
        Entry(f1, textvariable=self.input_file, width=40).grid(row=0, column=1)
        Button(f1, text="Abrir", command=self.browse_input).grid(row=0, column=2)

        f2 = Frame(root); f2.pack(padx=20, pady=10, fill="x")
        Label(f2, text="Pasta Destino:").grid(row=0, column=0)
        Entry(f2, textvariable=self.output_dir, width=40).grid(row=0, column=1)
        Button(f2, text="Pasta", command=self.browse_output).grid(row=0, column=2)

        f3 = Frame(root); f3.pack(padx=20, pady=10, fill="x")
        Radiobutton(f3, text="Qtd Arquivos", variable=self.split_mode, value=1).grid(row=0, column=0)
        Entry(f3, textvariable=self.num_files).grid(row=0, column=1)
        Radiobutton(f3, text="Linhas/Arq", variable=self.split_mode, value=2).grid(row=1, column=0)
        Entry(f3, textvariable=self.rows_per_file).grid(row=1, column=1)
        
        Label(f3, text="Formato Saída:").grid(row=2, column=0)
        ttk.Combobox(f3, textvariable=self.output_format, values=["csv", "xlsx"]).grid(row=2, column=1)

        self.progress_label = Label(root, text="")
        self.progress_label.pack()
        
        self.btn = Button(root, text="PROCESSAR", command=self.start, bg="#27ae60", fg="white")
        self.btn.pack(pady=20)

    def browse_input(self):
        f = filedialog.askopenfilename(filetypes=[("Planilhas", "*.csv *.xlsx *.xls")])
        if f: self.input_file.set(f)

    def browse_output(self):
        d = filedialog.askdirectory()
        if d: self.output_dir.set(d)

    def start(self):
        threading.Thread(target=self.run, daemon=True).start()

    def run(self):
        try:
            # AQUI ESTÁ A CORREÇÃO: O nome do parâmetro deve bater com a função
            res = split_file(
                input_path=self.input_file.get(), # Mudei de input_csv para input_path
                output_dir=self.output_dir.get(),
                output_format=self.output_format.get(),
                num_files=self.num_files.get() if self.split_mode.get()==1 else None,
                rows_per_file=self.rows_per_file.get() if self.split_mode.get()==2 else None,
                progress_callback=lambda m: self.progress_label.config(text=m)
            )
            
            zip_path = Path(self.output_dir.get()) / "resultado_split.zip"
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for f in res: zipf.write(f, arcname=Path(f).name)
            messagebox.showinfo("Sucesso", "Concluído!")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

if __name__ == "__main__":
    root = Tk()
    CSVSplitterApp(root)
    root.mainloop()